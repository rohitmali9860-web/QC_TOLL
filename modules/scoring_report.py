"""
scoring_report.py - Module 6: Staged Scoring & Final Report Engine
Calculates sector-by-sector scores and rolls up into an overall percentage.
Enforces the 50% Quality Gate (<50% locked + correction checklist, >=50% unlocked).
Generates comprehensive ReportLab PDF QC Reports with embedded screenshot attachments
and Excel audit workbooks.
"""

import os
import io
import openpyxl
from datetime import datetime
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image as RLImage, KeepTogether, PageBreak, HRFlowable
)

class StagedScoringEngine:
    def __init__(self, pass_threshold=50.0):
        self.pass_threshold = pass_threshold

    def calculate_overall_score(self, sector_results):
        """
        Calculates rolled-up overall score across all completed sectors.
        """
        total_checks = 0
        passed_checks = 0
        failed_checks = 0
        sectors_summary = []

        for sec in sector_results:
            sec_name = sec.get('sector_name', 'Sector')
            sec_score = float(sec.get('score', 0.0))
            sec_total = int(sec.get('total_checks', 0))
            sec_passed = int(sec.get('passed_checks', 0))
            sec_failed = int(sec.get('failed_checks', 0))
            status = sec.get('status', 'FAIL')

            total_checks += sec_total
            passed_checks += sec_passed
            failed_checks += sec_failed

            sectors_summary.append({
                'sector_name': sec_name,
                'score': sec_score,
                'total_checks': sec_total,
                'passed_checks': sec_passed,
                'failed_checks': sec_failed,
                'status': status
            })

        overall_score = round((passed_checks / max(total_checks, 1)) * 100.0, 1)
        passed_gate = overall_score >= self.pass_threshold

        # Aggregate all failed checks into an actionable correction checklist
        correction_checklist = []
        for sec in sector_results:
            for chk in sec.get('checks', []):
                if chk.get('status') == 'FAIL':
                    correction_checklist.append({
                        'sector_name': sec.get('sector_name', 'QC Sector'),
                        'field_name': chk.get('field_name', 'Unknown Field'),
                        'expected': chk.get('expected', 'N/A'),
                        'actual': chk.get('actual', 'N/A'),
                        'severity': chk.get('severity', 'HIGH'),
                        'details': chk.get('details', 'Discrepancy detected.'),
                        'bartender_fix': self._generate_bartender_instruction(chk)
                    })

        return {
            'overall_score': overall_score,
            'pass_threshold': self.pass_threshold,
            'passed_gate': passed_gate,
            'status': 'PASS' if passed_gate else 'FAILED - RE-TEST REQUIRED',
            'total_checks': total_checks,
            'passed_checks': passed_checks,
            'failed_checks': failed_checks,
            'sectors': sectors_summary,
            'correction_checklist': correction_checklist
        }

    def _generate_bartender_instruction(self, check):
        """Generates actionable guidance for the designer to resolve errors in BarTender."""
        field = check.get('field_name', '')
        details = check.get('details', '')

        if 'font' in field.lower() or 'font' in details.lower():
            return f"Open BarTender template (.btw) -> Select '{field}' object -> Properties -> Font -> Update font family/size to match approved artwork specification."
        elif 'placement' in field.lower() or 'margin' in details.lower():
            return f"Open BarTender template -> Select '{field}' -> Transform -> Adjust X/Y coordinates to remain within the safe print boundary (>1.0mm margin)."
        elif 'missing' in details.lower() or 'presence' in field.lower():
            return f"Open BarTender template -> Add missing text/barcode object for '{field}' -> Link data source to database column."
        elif 'rfid' in field.lower() or 'epc' in details.lower():
            return f"Check RFID encoder settings in BarTender -> Ensure SGTIN-96 format is selected with valid 24-character hexadecimal data source."
        elif 'mapping' in details.lower() or 'data mismatch' in details.lower():
            return f"Verify Database Connection Setup in BarTender -> Ensure layout field '{field}' is bound to the correct Order CSV/Excel column."
        else:
            return f"Inspect '{field}' in BarTender template and adjust to align with: {check.get('expected')}."

    def generate_pdf_report(self, run_metadata, sector_results, screenshots=None, output_path=None):
        """
        Generates executive quality control audit PDF report using ReportLab.
        Embeds metadata, score badges, pass/fail matrices, and user-attached screenshots.
        """
        score_rollup = self.calculate_overall_score(sector_results)
        if not score_rollup['passed_gate']:
            raise ValueError(f"Cannot generate Final QC PDF Report: Overall score {score_rollup['overall_score']}% is below the {self.pass_threshold}% pass threshold.")

        buf = io.BytesIO()
        target = output_path or buf
        doc = SimpleDocTemplate(
            target,
            pagesize=letter,
            rightMargin=36,
            leftMargin=36,
            topMargin=36,
            bottomMargin=36
        )

        styles = getSampleStyleSheet()
        
        # Custom typography styles
        title_style = ParagraphStyle(
            'DocTitle',
            parent=styles['Normal'],
            fontName='Helvetica-Bold',
            fontSize=18,
            leading=22,
            textColor=colors.HexColor('#0D2363')
        )
        subtitle_style = ParagraphStyle(
            'DocSubtitle',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=10,
            leading=14,
            textColor=colors.HexColor('#475569')
        )
        h2_style = ParagraphStyle(
            'Heading2',
            parent=styles['Normal'],
            fontName='Helvetica-Bold',
            fontSize=12,
            leading=16,
            textColor=colors.HexColor('#0D2363'),
            spaceBefore=12,
            spaceAfter=6
        )
        cell_style = ParagraphStyle(
            'TableCell',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=8,
            leading=10,
            textColor=colors.HexColor('#1E293B')
        )
        cell_bold = ParagraphStyle(
            'TableCellBold',
            parent=styles['Normal'],
            fontName='Helvetica-Bold',
            fontSize=8,
            leading=10,
            textColor=colors.HexColor('#1E293B')
        )

        elements = []

        # 1. Header Banner & Title
        elements.append(Paragraph("<b>r-pac International</b> — BarTender Layout QC Audit Certificate", subtitle_style))
        elements.append(Paragraph("PRE-PRODUCTION QUALITY CONTROL VERIFICATION REPORT", title_style))
        elements.append(Spacer(1, 10))
        elements.append(HRFlowable(width="100%", thickness=1.5, color=colors.HexColor('#0D2363'), spaceAfter=12))

        # 2. Metadata Table & Score Badge
        rpo = run_metadata.get('rpo_number', '1000341139')
        designer = run_metadata.get('designer_name', 'Rohit')
        item = run_metadata.get('item_code', 'LS-SS26-PT')
        timestamp = run_metadata.get('timestamp', datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        overall_score = score_rollup['overall_score']

        meta_data = [
            [
                Paragraph("<b>Active RPO:</b>", cell_bold), Paragraph(str(rpo), cell_style),
                Paragraph("<b>Overall Score:</b>", cell_bold), Paragraph(f"<font color='#16A34A'><b>{overall_score}% (PASSED)</b></font>", cell_bold)
            ],
            [
                Paragraph("<b>Designer:</b>", cell_bold), Paragraph(str(designer), cell_style),
                Paragraph("<b>Pass Threshold:</b>", cell_bold), Paragraph(f"{self.pass_threshold}%", cell_style)
            ],
            [
                Paragraph("<b>Item Code:</b>", cell_bold), Paragraph(str(item), cell_style),
                Paragraph("<b>Total Checks:</b>", cell_bold), Paragraph(f"{score_rollup['passed_checks']} Passed / {score_rollup['failed_checks']} Failed", cell_style)
            ],
            [
                Paragraph("<b>Audit Date:</b>", cell_bold), Paragraph(str(timestamp), cell_style),
                Paragraph("<b>Final Status:</b>", cell_bold), Paragraph("<font color='#16A34A'><b>APPROVED FOR PRODUCTION</b></font>", cell_bold)
            ]
        ]
        t_meta = Table(meta_data, colWidths=[80, 180, 90, 190])
        t_meta.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#F8FAFC')),
            ('BOX', (0,0), (-1,-1), 1.0, colors.HexColor('#CBD5E1')),
            ('INNERGRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E2E8F0')),
            ('PADDING', (0,0), (-1,-1), 5),
        ]))
        elements.append(t_meta)
        elements.append(Spacer(1, 14))

        # 3. Sector Score Breakdown Table
        elements.append(Paragraph("1. Staged Sector Score Summary", h2_style))
        sector_headers = [Paragraph("<b>QC Sector</b>", cell_bold), Paragraph("<b>Total Checks</b>", cell_bold), Paragraph("<b>Passed</b>", cell_bold), Paragraph("<b>Failed</b>", cell_bold), Paragraph("<b>Score %</b>", cell_bold), Paragraph("<b>Status</b>", cell_bold)]
        sector_rows = [sector_headers]

        for s in score_rollup['sectors']:
            st_color = "#16A34A" if s['score'] >= 50.0 else "#DC2626"
            sector_rows.append([
                Paragraph(s['sector_name'], cell_style),
                Paragraph(str(s['total_checks']), cell_style),
                Paragraph(str(s['passed_checks']), cell_style),
                Paragraph(str(s['failed_checks']), cell_style),
                Paragraph(f"<b>{s['score']}%</b>", cell_bold),
                Paragraph(f"<font color='{st_color}'><b>{s['status']}</b></font>", cell_bold)
            ])

        t_sec = Table(sector_rows, colWidths=[180, 70, 60, 60, 80, 90])
        t_sec.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#0D2363')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('BOTTOMPADDING', (0,0), (-1,0), 6),
            ('TOPPADDING', (0,0), (-1,0), 6),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CBD5E1')),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F8FAFC')])
        ]))
        elements.append(t_sec)
        elements.append(Spacer(1, 14))

        # 4. Detailed Field-by-Field Findings & Remaining Issues
        elements.append(Paragraph("2. Detailed Findings & Verification Log", h2_style))
        finding_headers = [Paragraph("<b>Sector</b>", cell_bold), Paragraph("<b>Field / Check</b>", cell_bold), Paragraph("<b>Expected Spec</b>", cell_bold), Paragraph("<b>Actual Layout Finding</b>", cell_bold), Paragraph("<b>Result</b>", cell_bold)]
        finding_rows = [finding_headers]

        for sec in sector_results:
            for chk in sec.get('checks', []):
                is_pass = chk.get('status') == 'PASS'
                res_html = "<font color='#16A34A'>PASS</font>" if is_pass else "<font color='#DC2626'>FAIL</font>"
                finding_rows.append([
                    Paragraph(sec.get('sector_name', '')[:20], cell_style),
                    Paragraph(chk.get('field_name', ''), cell_style),
                    Paragraph(str(chk.get('expected', ''))[:45], cell_style),
                    Paragraph(str(chk.get('actual', ''))[:45], cell_style),
                    Paragraph(res_html, cell_bold)
                ])

        t_find = Table(finding_rows[:35], colWidths=[110, 120, 140, 120, 50])
        t_find.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1E293B')),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E2E8F0')),
            ('PADDING', (0,0), (-1,-1), 4),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F8FAFC')])
        ]))
        elements.append(t_find)
        elements.append(Spacer(1, 14))

        # 5. Embedded Problem Screenshots Section
        if screenshots:
            elements.append(Paragraph("3. Attached Problem Screenshots & Visual Evidence", h2_style))
            for s in screenshots:
                path = s.get('file_path')
                caption = s.get('caption', 'Defect visual proof')
                if path and os.path.exists(path):
                    elements.append(Spacer(1, 4))
                    elements.append(Paragraph(f"<b>Screenshot:</b> {caption}", cell_bold))
                    elements.append(Spacer(1, 4))
                    try:
                        rl_img = RLImage(path, width=400, height=220)
                        elements.append(rl_img)
                        elements.append(Spacer(1, 8))
                    except Exception:
                        elements.append(Paragraph(f"[Image could not be rendered: {path}]", cell_style))

        # 6. Signoff Footer
        elements.append(Spacer(1, 10))
        elements.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor('#CBD5E1'), spaceAfter=8))
        elements.append(Paragraph("Quality Assurance Sign-Off: Approved for automated BarTender spool generation. Certificate generated by r-pac QC Suite.", subtitle_style))

        doc.build(elements)

        if output_path:
            return output_path
        else:
            buf.seek(0)
            return buf

    def generate_excel_report(self, run_metadata, sector_results, output_path=None):
        """Generates comprehensive Excel audit report with multiple tabs."""
        wb = openpyxl.Workbook()
        
        # Tab 1: Executive Summary
        ws_summary = wb.active
        ws_summary.title = "Executive Summary"

        score_rollup = self.calculate_overall_score(sector_results)

        ws_summary.append(["r-pac International - BarTender Layout QC Audit Report"])
        ws_summary.append(["Generated At", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
        ws_summary.append(["Designer", run_metadata.get('designer_name', 'Rohit')])
        ws_summary.append(["Active RPO", run_metadata.get('rpo_number', '1000341139')])
        ws_summary.append(["Item Code", run_metadata.get('item_code', 'LS-SS26-PT')])
        ws_summary.append(["Overall Score", f"{score_rollup['overall_score']}%"])
        ws_summary.append(["Status", score_rollup['status']])
        ws_summary.append([])

        ws_summary.append(["Sector Name", "Score %", "Total Checks", "Passed", "Failed", "Status"])
        for s in score_rollup['sectors']:
            ws_summary.append([s['sector_name'], s['score'], s['total_checks'], s['passed_checks'], s['failed_checks'], s['status']])

        # Tab 2: Detailed Checks
        ws_details = wb.create_sheet(title="Detailed Checks")
        ws_details.append(["Sector", "Field Name", "Expected Spec", "Actual Layout Value", "Status", "Severity", "Details"])

        for sec in sector_results:
            for chk in sec.get('checks', []):
                ws_details.append([
                    sec.get('sector_name', ''),
                    chk.get('field_name', ''),
                    chk.get('expected', ''),
                    chk.get('actual', ''),
                    chk.get('status', ''),
                    chk.get('severity', ''),
                    chk.get('details', '')
                ])

        # Style Worksheets
        for ws in [ws_summary, ws_details]:
            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = openpyxl.utils.get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len + 3, 14)

        if output_path:
            wb.save(output_path)
            return output_path
        else:
            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            return buf
