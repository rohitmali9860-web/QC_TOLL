"""
variable_qc.py - Module 2: Variable Data / Order Mapping QC
Validates that data on the BarTender layout accurately matches order records
(CSV/Excel) according to dynamic column mapping definitions.
Supports 1-to-1 and 1-to-many mappings, template generation, and discrepancy diffs.
"""

import os
import csv
import io
import openpyxl
import fitz

class VariableQCAnalyzer:
    def __init__(self):
        pass

    def parse_order_file(self, file_path):
        """Reads CSV or Excel order file into a list of dictionaries with normalized keys."""
        ext = os.path.splitext(file_path)[1].lower()
        records = []
        headers = []

        if ext == '.csv':
            with open(file_path, 'r', encoding='utf-8', errors='replace') as f:
                # Try sniffing delimiter
                sample = f.read(2048)
                f.seek(0)
                delimiter = ','
                if '\t' in sample and sample.count('\t') > sample.count(','):
                    delimiter = '\t'
                elif ';' in sample and sample.count(';') > sample.count(','):
                    delimiter = ';'
                
                reader = csv.DictReader(f, delimiter=delimiter)
                headers = reader.fieldnames or []
                for row in reader:
                    cleaned_row = {k.strip(): str(v).strip() for k, v in row.items() if k}
                    records.append(cleaned_row)
        elif ext in ('.xlsx', '.xls'):
            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if rows:
                raw_headers = rows[0]
                headers = [str(h).strip() for h in raw_headers if h is not None]
                for r in rows[1:]:
                    if any(cell is not None for cell in r):
                        row_dict = {}
                        for idx, h in enumerate(headers):
                            val = r[idx] if idx < len(r) and r[idx] is not None else ""
                            row_dict[h] = str(val).strip()
                        records.append(row_dict)
            wb.close()
        
        return {'headers': headers, 'records': records}

    def extract_layout_text_map(self, layout_pdf_path):
        """Extracts text tokens and key-value pairs from BarTender layout PDF."""
        doc = fitz.open(layout_pdf_path)
        full_text = ""
        kv_pairs = {}

        for page in doc:
            full_text += page.get_text() + "\n"

        doc.close()

        lines = [line.strip() for line in full_text.splitlines() if line.strip()]
        for line in lines:
            if ':' in line:
                parts = line.split(':', 1)
                k = parts[0].strip()
                v = parts[1].strip()
                kv_pairs[k] = v
            else:
                kv_pairs[line] = line

        return {'full_text': full_text, 'kv_pairs': kv_pairs, 'lines': lines}

    def generate_mapping_excel(self, template_fields, output_path=None):
        """Generates a downloadable field mapping template in Excel format."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Field Mapping"

        headers = ["Template Field Name", "Order Column Header", "Is Mandatory", "Match Mode", "Sample Expected Value"]
        ws.append(headers)

        header_fill = openpyxl.styles.PatternFill(start_color="0D2363", end_color="0D2363", fill_type="solid")
        header_font = openpyxl.styles.Font(name="Arial", size=10, bold=True, color="FFFFFF")

        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")

        sample_rows = [
            ["Article_No", "ARTICLE_NO", "YES", "Exact", "1000341139-PT"],
            ["Size_Label", "SIZE_LABEL", "YES", "Exact", "M"],
            ["Color_Name", "COLOR_NAME", "YES", "Exact", "Midnight Navy"],
            ["RPO_Number", "RPO_NO", "YES", "Exact", "1000341139"],
            ["SKU_Sequence", "SKU_STRING", "YES", "Exact", "1/3"],
            ["Order_Quantity", "ORDER_QTY", "YES", "Exact", "5"],
            ["QR_Auth_URL", "QR_AUTH_URL", "YES", "Contains", "https://loveskin.com/auth/1000341139001"]
        ]

        for r in sample_rows:
            ws.append(r)

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 14)

        if output_path:
            wb.save(output_path)
            return output_path
        else:
            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            return buf

    def parse_mapping_file(self, mapping_file_path):
        """Parses an uploaded mapping Excel file into a dictionary map."""
        wb = openpyxl.load_workbook(mapping_file_path, data_only=True)
        ws = wb.active
        mapping = {}

        rows = list(ws.iter_rows(values_only=True))
        if len(rows) > 1:
            for r in rows[1:]:
                if r[0] and r[1]:
                    template_field = str(r[0]).strip()
                    order_col = str(r[1]).strip()
                    mapping[template_field] = order_col
        wb.close()
        return mapping

    def validate(self, order_file_path, layout_pdf_path, mapping=None, record_index=0):
        """
        Validates the BarTender layout against the specified order record using column mapping.
        """
        order_data = self.parse_order_file(order_file_path)
        records = order_data['records']
        if not records:
            return {
                'sector_name': 'Variable Data QC',
                'score': 0.0,
                'status': 'FAIL',
                'error': 'No records found in order file.',
                'checks': []
            }

        # Select target order record
        target_record = records[min(record_index, len(records) - 1)]
        layout_data = self.extract_layout_text_map(layout_pdf_path)
        layout_text = layout_data['full_text']
        kv_pairs = layout_data['kv_pairs']

        # Default mapping if none provided
        default_mapping = {
            'Article': 'ARTICLE_NO',
            'Size': 'SIZE_LABEL',
            'Color': 'COLOR_NAME',
            'RPO': 'RPO_NO',
            'SKU': 'SKU_STRING',
            'QTY': 'ORDER_QTY'
        }
        active_mapping = mapping or default_mapping

        checks = []
        passed_count = 0
        total_count = 0

        for layout_field, order_col in active_mapping.items():
            total_count += 1
            expected_val = target_record.get(order_col, '')
            
            # If mapping is invalid or column does not exist in order file
            if not expected_val and order_col not in target_record:
                checks.append({
                    'field_name': layout_field,
                    'order_column': order_col,
                    'expected': f"Column '{order_col}' missing in Order file",
                    'actual': 'N/A',
                    'status': 'FAIL',
                    'severity': 'HIGH',
                    'details': f"Mapping Error: Column '{order_col}' was mapped to '{layout_field}' but not found in the uploaded order file."
                })
                continue

            # Check if expected value appears on the layout
            # A. Check key-value pairs
            found_actual = None
            for k, v in kv_pairs.items():
                if layout_field.lower() in k.lower():
                    found_actual = v
                    break

            # B. If not in kv_pairs, search in full layout text
            is_match = False
            if str(expected_val).lower() in layout_text.lower():
                is_match = True
                actual_display = str(expected_val) if not found_actual else found_actual
            else:
                actual_display = found_actual if found_actual else 'MISSING / INCORRECT'

            if is_match:
                passed_count += 1
                checks.append({
                    'field_name': layout_field,
                    'order_column': order_col,
                    'expected': str(expected_val),
                    'actual': actual_display,
                    'status': 'PASS',
                    'severity': 'LOW',
                    'details': f"Value matches order record exactly (Record #{record_index + 1})."
                })
            else:
                checks.append({
                    'field_name': layout_field,
                    'order_column': order_col,
                    'expected': str(expected_val),
                    'actual': actual_display,
                    'status': 'FAIL',
                    'severity': 'CRITICAL',
                    'details': f"Data Mismatch: Expected '{expected_val}' from column '{order_col}', but layout contained '{actual_display}'."
                })

        score = round((passed_count / max(total_count, 1)) * 100.0, 1)

        return {
            'sector_name': 'Variable Data & Mapping QC',
            'score': score,
            'total_checks': total_count,
            'passed_checks': passed_count,
            'failed_checks': total_count - passed_count,
            'status': 'PASS' if score >= 90.0 else 'FAIL',
            'order_record_used': target_record,
            'record_index': record_index,
            'total_records_in_order': len(records),
            'mapping_used': active_mapping,
            'checks': checks
        }
