"""
sample_data_generator.py - Generates rich, realistic sample test assets for all QC modules:
- Signed Customer Artwork PDF
- BarTender PP Layout PDFs (Pass, Font Mismatch, Placement Shift, Missing Field)
- Order Data (CSV and XLSX)
- Mapping Template (XLSX)
- Thermal RFID ZPL Spool
- Problem Screenshots
"""

import os
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import mm
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from PIL import Image, ImageDraw, ImageFont

DATA_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'sample_data')

def ensure_dirs():
    os.makedirs(DATA_DIR, exist_ok=True)
    os.makedirs(os.path.join(DATA_DIR, 'screenshots'), exist_ok=True)
    os.makedirs(os.path.join(DATA_DIR, 'batch_layouts'), exist_ok=True)

from reportlab.lib.pagesizes import A4

def generate_signed_artwork():
    """Generates the Customer Signed Artwork PDF matching the exact r-pac Loveskin Pocket Tag specification sheet."""
    pdf_path = os.path.join(DATA_DIR, 'artwork_signed_spec.pdf')
    c = canvas.Canvas(pdf_path, pagesize=A4)
    page_w, page_h = A4

    # 1. Header Grid Table (Corporate Black & White)
    c.setStrokeColor(colors.black)
    c.setLineWidth(0.75)
    
    # Title banner
    c.setFillColor(colors.black)
    c.rect(10 * mm, page_h - 18 * mm, 190 * mm, 8 * mm, fill=1, stroke=1)
    c.setFillColor(colors.white)
    c.setFont("Helvetica-Bold", 10)
    c.drawString(13 * mm, page_h - 13.5 * mm, "PRODUCT CATEGORY: POCKET TAG")

    # Header parameters box
    c.setFillColor(colors.white)
    c.rect(10 * mm, page_h - 40 * mm, 190 * mm, 22 * mm, fill=0, stroke=1)
    c.line(75 * mm, page_h - 18 * mm, 75 * mm, page_h - 40 * mm)
    c.line(140 * mm, page_h - 18 * mm, 140 * mm, page_h - 40 * mm)
    c.line(10 * mm, page_h - 29 * mm, 190 * mm, page_h - 29 * mm)

    c.setFillColor(colors.black)
    c.setFont("Helvetica-Bold", 7.5)
    c.drawString(12 * mm, page_h - 23 * mm, "DATE: 24/06/2026")
    c.drawString(12 * mm, page_h - 34 * mm, "ITEM REF: LOVESKIN-PT")

    c.drawString(77 * mm, page_h - 23 * mm, "CUSTOMER: LOVESKIN")
    c.drawString(77 * mm, page_h - 34 * mm, "SIZE FINISHED: 25 mm x 60 mm")

    c.drawString(142 * mm, page_h - 23 * mm, "MATERIAL: N/A")
    c.drawString(142 * mm, page_h - 34 * mm, "RFID INLAY: R-BOOST M830")

    # 2. Dieline Dimension Box (Front & Back Views)
    # Dimensions: 60mm width x 25mm height
    front_x = 40 * mm
    front_y = page_h - 100 * mm
    back_x = 120 * mm
    back_y = page_h - 100 * mm
    tag_w = 60 * mm
    tag_h = 25 * mm

    # Dieline dimensions text
    c.setFont("Helvetica-Bold", 7.5)
    c.setFillColor(colors.HexColor("#dc2626"))
    c.drawCentredString(front_x + tag_w/2, front_y + tag_h + 3*mm, "60 mm")
    c.drawRightString(front_x - 3*mm, front_y + tag_h/2, "25 mm")
    c.drawCentredString(back_x + tag_w - 5*mm, back_y + tag_h + 3*mm, "3 x 6 mm Black Mark")

    # Front Tag (Inlay View)
    c.setStrokeColor(colors.HexColor("#38bdf8"))
    c.setLineWidth(1.0)
    c.rect(front_x, front_y, tag_w, tag_h)
    c.setFillColor(colors.HexColor("#64748b"))
    c.setFont("Helvetica", 7.5)
    c.drawCentredString(front_x + tag_w/2, front_y - 5*mm, "Front")

    # Inlay outline inside front tag (42mm x 16mm)
    inlay_w = 42 * mm
    inlay_h = 16 * mm
    c.setStrokeColor(colors.HexColor("#94a3b8"))
    c.rect(front_x + (tag_w - inlay_w)/2, front_y + (tag_h - inlay_h)/2, inlay_w, inlay_h)
    c.drawCentredString(front_x + tag_w/2, front_y + tag_h/2 - 2*mm, "R-BOOST M830 (42x16mm)")

    # Back Tag (Production Visual Artwork)
    c.setStrokeColor(colors.HexColor("#38bdf8"))
    c.rect(back_x, back_y, tag_w, tag_h)
    c.setFillColor(colors.HexColor("#64748b"))
    c.drawCentredString(back_x + tag_w/2, back_y - 5*mm, "Back")

    # Black Mark (3 x 6 mm) at top right
    c.setFillColor(colors.black)
    c.rect(back_x + tag_w - 6*mm, back_y + tag_h - 3*mm, 6*mm, 3*mm, fill=1, stroke=0)

    # Back Tag Content
    c.setFont("Helvetica-Bold", 8)
    c.setFillColor(colors.HexColor("#0f172a"))
    c.drawString(back_x + 3*mm, back_y + 16*mm, "LoveSkin")

    c.setFont("Helvetica", 5.0)
    c.drawString(back_x + 3*mm, back_y + 11*mm, "Taglia / Size: S")
    c.drawString(back_x + 3*mm, back_y + 7.5*mm, "Articolo / Article: Z5E213TS")
    c.drawString(back_x + 3*mm, back_y + 4*mm, "Colore / Color: Nero")

    # RFID Logo
    c.setFont("Helvetica-Bold", 5)
    c.drawString(back_x + 3*mm, back_y + 1*mm, "[RFID]")

    # QR Code placeholder (15x15mm)
    qr_x = back_x + tag_w - 22*mm
    qr_y = back_y + 3*mm
    c.setFont("Helvetica-Bold", 4.5)
    c.drawCentredString(qr_x + 7.5*mm, qr_y + 16*mm, "QR CODE")
    c.rect(qr_x, qr_y, 15*mm, 15*mm, fill=0, stroke=1)
    c.drawCentredString(qr_x + 7.5*mm, qr_y + 7*mm, "[QR 15x15mm]")

    # 3. 10-Row Spec Field Table Matrix (Bottom Right)
    table_x = 100 * mm
    table_y = page_h - 180 * mm
    table_w = 100 * mm

    c.setFont("Helvetica-Bold", 7)
    c.setFillColor(colors.HexColor("#dc2626"))
    c.drawString(table_x, table_y + 32*mm, "SPECIFICATION FIELD MATRIX (1-10):")

    # Table Header
    c.setFillColor(colors.HexColor("#f8fafc"))
    c.rect(table_x, table_y + 24*mm, table_w, 6*mm, fill=1, stroke=1)
    c.setFillColor(colors.black)
    c.setFont("Helvetica-Bold", 6)
    c.drawString(table_x + 2*mm, table_y + 26*mm, "#  Description")
    c.drawString(table_x + 38*mm, table_y + 26*mm, "Font")
    c.drawString(table_x + 72*mm, table_y + 26*mm, "Field info")

    spec_rows = [
        ("1", "Brand logo", "Brand Spec", "Mandatory/ Fixed"),
        ("2", "Size", "Arial Medium - 5pt", "Mandatory/ Fixed"),
        ("3", "Article number", "Arial Medium - 5pt", "Mandatory/ Fixed"),
        ("4", "Color", "Arial Medium - 5pt", "Mandatory/ Fixed"),
        ("5", "RFID Logo", "ISO Symbol", "Mandatory/ Fixed"),
        ("6", "Size (max 12 digits)", "Arial Medium - 5pt", "Mandatory/ Variable"),
        ("7", "Article number (12 digits)", "Arial Medium - 5pt", "Mandatory/ Variable"),
        ("8", "Color (15 digits)", "Arial Medium - 5pt", "Mandatory/ Variable"),
        ("9", "QR Code Text", "Arial Medium - 5pt", "Mandatory/ Fixed"),
        ("10", "QR Code (15 x15mm)", "2D Barcode", "Mandatory/ Variable")
    ]

    c.setFont("Helvetica", 5.5)
    row_y = table_y + 20 * mm
    for r in spec_rows:
        c.rect(table_x, row_y, table_w, 4*mm, fill=0, stroke=1)
        c.drawString(table_x + 2*mm, row_y + 1.2*mm, f"{r[0]}. {r[1]}")
        c.drawString(table_x + 38*mm, row_y + 1.2*mm, r[2])
        c.drawString(table_x + 72*mm, row_y + 1.2*mm, r[3])
        row_y -= 4*mm

    # Inlay Spec Box (Bottom Left)
    c.setFont("Helvetica-Bold", 7.5)
    c.setFillColor(colors.HexColor("#0f172a"))
    c.drawString(20 * mm, page_h - 155 * mm, "Inlay: R-BOOST M830")
    c.setFont("Helvetica", 6.5)
    c.drawString(20 * mm, page_h - 160 * mm, "42 x 16 mm – 128 bits")

    c.save()
    print("Generated authentic A4 Loveskin Customer Signed Artwork Spec PDF.")
    return pdf_path

def generate_layout_pdf(filename, variant='pass'):
    """Generates BarTender Production Proof (PP PDF) layout variants matching Loveskin 60mm x 25mm Pocket Tag."""
    pdf_path = os.path.join(DATA_DIR, filename)
    # 60mm width x 25mm height pocket tag
    w = 60 * mm
    h = 25 * mm
    c = canvas.Canvas(pdf_path, pagesize=(w, h))

    # Cut line / Dieline border
    c.setStrokeColor(colors.HexColor("#0f172a"))
    c.setLineWidth(0.8)
    c.rect(0.2*mm, 0.2*mm, w - 0.4*mm, h - 0.4*mm)

    if variant == 'pass':
        # Black Mark (3 x 6 mm) at top right
        c.setFillColor(colors.black)
        c.rect(w - 6.2*mm, h - 3.2*mm, 6*mm, 3*mm, fill=1, stroke=0)

        # 1. Brand Logo (Fixed, matching spec)
        c.setFont("Helvetica-Bold", 8)
        c.setFillColor(colors.HexColor("#0f172a"))
        c.drawString(3 * mm, 16 * mm, "LoveSkin")

        # 2-4 & 6-8. Fixed Labels & Variable Values (Arial / Helvetica 5.0pt)
        c.setFont("Helvetica", 5.0)
        c.drawString(3 * mm, 11 * mm, "Taglia / Size: S")
        c.drawString(3 * mm, 7.5 * mm, "Articolo / Article: Z5E213TS")
        c.drawString(3 * mm, 4 * mm, "Colore / Color: Nero")

        # 5. RFID Logo
        c.setFont("Helvetica-Bold", 5.0)
        c.drawString(3 * mm, 1 * mm, "[RFID]")

        # 9-10. QR Code Text (5pt) and QR Code box (15x15mm)
        qr_x = w - 22 * mm
        qr_y = 3 * mm
        c.setFont("Helvetica-Bold", 4.5)
        c.drawCentredString(qr_x + 7.5*mm, qr_y + 16*mm, "QR CODE")
        c.rect(qr_x, qr_y, 15*mm, 15*mm, fill=0, stroke=1)
        c.drawCentredString(qr_x + 7.5*mm, qr_y + 7*mm, "[QR 15x15mm]")

    elif variant == 'fail_font_placement' or variant == 'fail':
        # Simulated font & size discrepancy (Times-Roman 7pt instead of Arial 5pt)
        c.setFont("Times-Roman", 7.0)
        c.drawString(3 * mm, 16 * mm, "LoveSkin")
        c.setFont("Courier", 6.5)
        c.drawString(3 * mm, 11 * mm, "Size: Small (Wrong Label)")
        c.drawString(3 * mm, 7.5 * mm, "Article: 1000341139 (Mismatch)")
        c.drawString(3 * mm, 4 * mm, "Color: Black (Mismatch)")

        # Missing QR Code text & shifted QR box
        qr_x = w - 28 * mm
        qr_y = 5 * mm
        c.rect(qr_x, qr_y, 12*mm, 12*mm, fill=0, stroke=1)

    elif variant == 'batch_sku2':
        # SKU 2 Pass
        c.setFont("Helvetica-Bold", 7.5)
        c.setFillColor(colors.black)
        c.drawCentredString(w/2, h - 7*mm, "LOVESKIN APPAREL")
        c.setFont("Helvetica-Bold", 6)
        c.setFillColor(colors.HexColor("#1e40af"))
        c.drawString(2*mm, h - 12*mm, "[((o))] RFID")
        c.setFont("Helvetica", 5.2)
        c.setFillColor(colors.black)
        c.drawString(2*mm, h - 17*mm, "Article: 1000341139-PT")
        c.drawString(2*mm, h - 21*mm, "Size: L")
        c.drawString(2*mm, h - 25*mm, "Color: Midnight Navy")
        c.drawString(2*mm, h - 29*mm, "RPO: 1000341139")
        c.drawString(2*mm, h - 33*mm, "SKU: 2/3")
        c.drawString(2*mm, h - 37*mm, "QTY: 5")
        c.setFillColor(colors.black)
        c.rect(2*mm, 12*mm, 21*mm, 7*mm, fill=1, stroke=0)
        c.setFont("Helvetica", 4.5)
        c.drawCentredString(w/2, 9*mm, "*1000341139002*")
        c.setLineWidth(0.5)
        c.rect(8*mm, 2*mm, 9*mm, 6*mm, fill=0, stroke=1)
        c.setFont("Helvetica", 4)
        c.drawCentredString(w/2, 4.5*mm, "https://loveskin.com/auth")

    elif variant == 'batch_sku3':
        # SKU 3 Pass
        c.setFont("Helvetica-Bold", 7.5)
        c.setFillColor(colors.black)
        c.drawCentredString(w/2, h - 7*mm, "LOVESKIN APPAREL")
        c.setFont("Helvetica-Bold", 6)
        c.setFillColor(colors.HexColor("#1e40af"))
        c.drawString(2*mm, h - 12*mm, "[((o))] RFID")
        c.setFont("Helvetica", 5.2)
        c.setFillColor(colors.black)
        c.drawString(2*mm, h - 17*mm, "Article: 1000341139-PT")
        c.drawString(2*mm, h - 21*mm, "Size: XL")
        c.drawString(2*mm, h - 25*mm, "Color: Midnight Navy")
        c.drawString(2*mm, h - 29*mm, "RPO: 1000341139")
        c.drawString(2*mm, h - 33*mm, "SKU: 3/3")
        c.drawString(2*mm, h - 37*mm, "QTY: 5")
        c.setFillColor(colors.black)
        c.rect(2*mm, 12*mm, 21*mm, 7*mm, fill=1, stroke=0)
        c.setFont("Helvetica", 4.5)
        c.drawCentredString(w/2, 9*mm, "*1000341139003*")
        c.setLineWidth(0.5)
        c.rect(8*mm, 2*mm, 9*mm, 6*mm, fill=0, stroke=1)
        c.setFont("Helvetica", 4)
        c.drawCentredString(w/2, 4.5*mm, "https://loveskin.com/auth")

    c.save()
    return pdf_path

def generate_order_files():
    """Generates realistic order CSV and XLSX files."""
    headers = [
        "RPO_NO", "SKU_INDEX", "SKU_TOTAL", "SKU_STRING", "ARTICLE_NO",
        "COLOR_NAME", "SIZE_LABEL", "ORDER_QTY", "EPC_HEADER_HEX",
        "COMPANY_PREFIX", "ITEM_REF", "SERIAL_START", "SERIAL_END",
        "SAMPLE_EPC_HEX", "QR_AUTH_URL"
    ]

    rows = [
        [
            "1000341139", 1, 3, "1/3", "Z5E213TS",
            "Nero", "S", 5, "30",
            "0843210", "100341", 1001, 1005,
            "3034027780000000000003E9", "https://loveskin.com/auth/1000341139001"
        ],
        [
            "1000341139", 2, 3, "2/3", "Z5E213TS",
            "Nero", "M", 5, "30",
            "0843210", "100341", 1006, 1010,
            "3034027780000000000003EE", "https://loveskin.com/auth/1000341139002"
        ],
        [
            "1000341139", 3, 3, "3/3", "Z5E213TS",
            "Nero", "L", 5, "30",
            "0843210", "100341", 1011, 1015,
            "3034027780000000000003F3", "https://loveskin.com/auth/1000341139003"
        ]
    ]

    # 1. CSV
    csv_path = os.path.join(DATA_DIR, '1000341139_Loveskin-PT_001.csv')
    with open(csv_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerows(rows)

    # 2. XLSX
    xlsx_path = os.path.join(DATA_DIR, '1000341139_Loveskin-PT_001.xlsx')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Order Data"

    # Styling
    header_fill = PatternFill(start_color="0D2363", end_color="0D2363", fill_type="solid")
    header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")

    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for r in rows:
        ws.append(r)

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    wb.save(xlsx_path)
    return csv_path, xlsx_path

def generate_mapping_template():
    """Generates standard BarTender layout field mapping Excel template."""
    xlsx_path = os.path.join(DATA_DIR, 'mapping_template_standard.xlsx')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Field Mapping"

    headers = ["Template Field Name", "Field Type", "Order File Column", "Is Mandatory", "Validation Rule / Regex", "Sample Output"]
    mappings = [
        ["Brand_Logo", "Fixed Text", "N/A (Fixed)", "YES", "Exact Match: LOVESKIN APPAREL", "LOVESKIN APPAREL"],
        ["RFID_Logo", "Fixed Icon", "N/A (Fixed)", "YES", "Exact Match: [((o))] RFID", "[((o))] RFID"],
        ["Article_No", "Variable Text", "ARTICLE_NO", "YES", "^[0-9]{10}-[A-Z]{2}$", "1000341139-PT"],
        ["Size_Label", "Variable Text", "SIZE_LABEL", "YES", "^(XS|S|M|L|XL|2XL|3XL|[0-9]{1,2})$", "M"],
        ["Color_Name", "Variable Text", "COLOR_NAME", "YES", "^[A-Za-z ]+$", "Midnight Navy"],
        ["RPO_Number", "Variable Text", "RPO_NO", "YES", "^[0-9]{10}$", "1000341139"],
        ["SKU_Sequence", "Variable Text", "SKU_STRING", "YES", "^[0-9]+/[0-9]+$", "1/3"],
        ["Order_Quantity", "Variable Text", "ORDER_QTY", "YES", "^[0-9]+$", "5"],
        ["EPC_Serial_Hex", "RFID Hex", "SAMPLE_EPC_HEX", "YES", "^30[0-9A-Fa-f]{22}$", "3034027780000000000003E9"],
        ["QR_Code_URL", "2D Barcode", "QR_AUTH_URL", "YES", "^https?://.+$", "https://loveskin.com/auth/1000341139001"]
    ]

    header_fill = PatternFill(start_color="0D2363", end_color="0D2363", fill_type="solid")
    header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")

    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for m in mappings:
        ws.append(m)

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 14)

    wb.save(xlsx_path)
    return xlsx_path

def generate_sample_zpl():
    """Generates sample RFID ZPL spool output."""
    zpl_path = os.path.join(DATA_DIR, 'sample_rfid_spool.zpl')
    zpl_content = """^XA
^PW400
^LL600
^PON
^LH0,0
^RS8,,,1
^RFW,H,,,^FD3034027780000000000003E9^FS
^FO60,40^A0N,28,28^FDLOVESKIN APPAREL^FS
^FO40,80^A0N,22,22^FD[((o))] RFID^FS
^FO40,120^A0N,18,18^FDArticle: 1000341139-PT^FS
^FO40,150^A0N,18,18^FDSize: M^FS
^FO40,180^A0N,18,18^FDColor: Midnight Navy^FS
^FO40,210^A0N,18,18^FDRPO: 1000341139^FS
^FO40,240^A0N,18,18^FDSKU: 1/3^FS
^FO40,270^A0N,18,18^FDQTY: 5^FS
^FO40,320^BY2,3,60^BCN,60,Y,N,N^FD1000341139001^FS
^FO140,420^BQN,2,4^FDQA,https://loveskin.com/auth/1000341139001^FS
^XZ
"""
    with open(zpl_path, 'w', encoding='utf-8') as f:
        f.write(zpl_content.strip())
    return zpl_path

def generate_sample_screenshots():
    """Generates sample defect screenshots for problem attachment demo."""
    screenshot_path1 = os.path.join(DATA_DIR, 'screenshots', 'defect_font_mismatch.png')
    img = Image.new('RGB', (600, 400), color='#f8fafc')
    draw = ImageDraw.Draw(img)

    # Draw simulated label fragment
    draw.rectangle([50, 50, 550, 350], fill='#ffffff', outline='#cbd5e1', width=2)
    draw.text((70, 70), "BarTender Layout Inspection - Problem Annotation", fill='#0f172a')
    draw.text((70, 110), "Brand Header: LOVESKIN APPAREL (Font: Times-Roman 9pt)", fill='#334155')
    draw.text((70, 150), "Expected Spec: Arial-Bold 7.5pt per Artwork", fill='#64748b')
    draw.text((70, 190), "Discrepancy: Mismatched font family & sizing exceeds ±0.5pt", fill='#dc2626')
    draw.text((70, 230), "Field X-Offset: 4.5mm (Spec: 2.0mm ±0.5mm)", fill='#dc2626')

    # Red circle callout
    draw.ellipse([60, 100, 500, 180], outline='#ef4444', width=3)
    draw.text((380, 80), "DEFECT #1: FONT MISMATCH", fill='#ef4444')

    img.save(screenshot_path1)
    return screenshot_path1

def build_all_sample_assets():
    ensure_dirs()
    p1 = generate_signed_artwork()
    p2 = generate_layout_pdf('layout_pass_sku1.pdf', 'pass')
    p3 = generate_layout_pdf('layout_fail_font_sku2.pdf', 'fail_font_placement')
    p4 = generate_layout_pdf(os.path.join('batch_layouts', '1000341139_SKU1_M.pdf'), 'pass')
    p5 = generate_layout_pdf(os.path.join('batch_layouts', '1000341139_SKU2_L.pdf'), 'batch_sku2')
    p6 = generate_layout_pdf(os.path.join('batch_layouts', '1000341139_SKU3_XL.pdf'), 'batch_sku3')
    c1, x1 = generate_order_files()
    m1 = generate_mapping_template()
    z1 = generate_sample_zpl()
    s1 = generate_sample_screenshots()
    print("All sample test assets generated successfully.")

if __name__ == '__main__':
    build_all_sample_assets()
